from flask import Blueprint, render_template, request, redirect, url_for, jsonify, make_response
from flask_login import login_required, current_user
from pymongo import MongoClient
import uuid
from datetime import datetime
from app.models.feedback_form.feedback_form import update_calculated_data
import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment, Font, Border, Side

# MongoDB Connection
client = MongoClient('mongodb://localhost:27017/')
db = client['faculty_feedback_db']
form_responses_collection = db["form_responses"]
calculated_data_collection = db["calculated_data"]

# Define ratings fields globally to avoid undefined variable warnings
ratings_fields = [
    "punctuality", "voice", "aids", "prep", "plan", "depth", "reaction",
    "behavior", "pace", "comm", "motiv", "expl", "clarity", "overall"
]
lab_ratings_fields = ["lab_expl", "lab_activ", "safety"]

feedback_bp = Blueprint('feedback_form', __name__)

@feedback_bp.route('/dashboard')
@login_required
def form_dashboard():
    user_forms = list(db.forms.find({'creator_id': current_user.id}).sort('modified_at', -1))
    for form in user_forms:
        if 'modified_at' in form and not isinstance(form['modified_at'], datetime):
            form['modified_at'] = datetime.strptime(form['modified_at'], '%Y-%m-%dT%H:%M:%S.%f')
    return render_template('feedback_form/form_dashboard.html', forms=user_forms)

@feedback_bp.route("/new")
@login_required
def new_form():
    form_id = str(uuid.uuid4())[:18]
    return redirect(url_for('feedback_form.edit_form', form_id=form_id))

@feedback_bp.route("/edit/<form_id>", methods=["GET", "POST"])
@login_required
def edit_form(form_id):
    if request.method == "GET":
        form = db.forms.find_one({'form_id': form_id, 'creator_id': current_user.id})
        if form:
            return render_template('feedback_form/edit_form.html', form=form)
        else:
            blank_form = {
                'form_id': form_id,
                'title': 'Untitled Form',
                'description': '',
                'subjects': []
            }
            return render_template('feedback_form/edit_form.html', form=blank_form)

    if request.method == "POST":
        if request.is_json:
            data = request.get_json()
            title = data.get('form_title', 'Untitled Form')
            description = data.get('form_description', '')
            subjects = data.get('subjects', [])
        else:
            title = request.form.get('title', 'Untitled Form')
            description = request.form.get('description', '')
            subjects = request.form.getlist('subjects[]')
        
        now = datetime.utcnow()
        form_data = {
            'form_id': form_id,
            'title': title,
            'description': description,
            'subjects': subjects,
            'creator_id': current_user.id,
            'modified_at': now
        }

        form = db.forms.find_one({'form_id': form_id, 'creator_id': current_user.id})
        if form:
            db.forms.update_one({'form_id': form_id}, {'$set': form_data})
        else:
            form_data['created_at'] = now
            db.forms.insert_one(form_data)

        return jsonify({"message": "Form saved successfully"})

@feedback_bp.route("/view/<form_id>")
def view_form(form_id):
    form = db.forms.find_one({'form_id': form_id})
    if form:
        return render_template('feedback_form/view_form.html', form=form)
    else:
        return 'Form not found', 404

@feedback_bp.route("/delete/<form_id>", methods=["POST"])
@login_required
def delete_form(form_id):
    result = db.forms.delete_one({'form_id': form_id, 'creator_id': current_user.id})
    if result.deleted_count:
        return jsonify({"message": "Form deleted successfully"})
    else:
        return jsonify({"message": "Form not found or deletion failed"}), 404

@feedback_bp.route("/submit/<form_id>", methods=["POST"])
def submit_form(form_id):
    form = db.forms.find_one({'form_id': form_id})
    if not form:
        return 'Form not found', 404

    raw_data = request.form.to_dict()

    structured_data = {
        'form_id': form_id,
        'submitted_at': datetime.utcnow(),
        'student_info': {},
        'subjects': [],
        'department_feedback': {}
    }

    student_fields = ['email', 'enrollment_number', 'student_name', 'discipline', 'branch', 'semester']
    for field in student_fields:
        structured_data['student_info'][field] = raw_data.get(field, '')

    subject_count = len(form['subjects'])
    for i in range(subject_count):
        section_num = i + 2
        subject_data = {
            'subject_name': form['subjects'][i]['subject_name'],
            'faculty_name': form['subjects'][i]['faculty_name'],
            'syllabus_covered': raw_data.get(f'syllabus_covered_{section_num}', ''),
            'ratings': {k: raw_data.get(f'{k}_{section_num}', '') for k in ratings_fields},
            'lab_ratings': {k: raw_data.get(f'{k}_{section_num}', '') for k in lab_ratings_fields}
        }
        structured_data['subjects'].append(subject_data)

    dept_fields = [
        'feedback_q1', 'improve_teaching', 'improve_department',
        'strengths_department', 'weaknesses_department',
        'strengths_university', 'weaknesses_university', 'suggestions_university'
    ]
    for field in dept_fields:
        structured_data['department_feedback'][field] = raw_data.get(field, '')

    db.form_responses.insert_one(structured_data)
    update_calculated_data(form_id)
    return redirect(url_for('feedback_form.submitted', form_id=form_id))

@feedback_bp.route("/submitted/<form_id>")
def submitted(form_id):
    return render_template('feedback_form/form_submitted.html', form_id=form_id)

@feedback_bp.route("/analysis/<form_id>")
@login_required
def form_analysis(form_id):
    form = db.forms.find_one({'form_id': form_id, 'creator_id': current_user.id})
    if not form:
        return 'Form not found or unauthorized', 403

    responses = list(db.form_responses.find({'form_id': form_id}))
    total_responses = len(responses)

    return render_template('analysis_form/responses_data.html', 
                          form=form, 
                          responses=responses, 
                          total_responses=total_responses,
                          segment='analysis')

@feedback_bp.route("/analysis/<form_id>/<subject_name>")
@login_required
def subject_analysis(form_id, subject_name):
    form = db.forms.find_one({'form_id': form_id, 'creator_id': current_user.id})
    if not form:
        return "Form not found or unauthorized", 403
    
    subject = next((s for s in form['subjects'] if s['subject_name'] == subject_name), None)
    if not subject:
        return "Subject not found", 404
    
    calculated_data = calculated_data_collection.find_one({"form_id": form_id, "subject_name": subject_name})
    if not calculated_data:
        return "No data calculated yet", 404

    return render_template('analysis_form/subject_analysis.html',
                          form=form,
                          subject=subject,
                          calculated_data=calculated_data,
                          segment='subject_analysis',
                          subject_name=subject_name)

@feedback_bp.route("/download_excel/<form_id>/<subject_name>", methods=["GET"])
@login_required
def download_excel(form_id, subject_name):
    # Check if form exists and belongs to the creator
    form = db.forms.find_one({'form_id': form_id, 'creator_id': current_user.id})
    if not form:
        return "Form not found or unauthorized", 403

    # Fetch all responses for this form
    responses = list(db.form_responses.find({'form_id': form_id}))

    # Fetch the stream (discipline), department (branch), and semester from the most recent response
    stream = "Unknown_Stream"  # Default in case no responses are found
    department_name = "Unknown_Department"
    semester = "Unknown_Semester"
    if responses:
        latest_response = responses[-1]  # Get the most recent response
        stream = latest_response.get('student_info', {}).get('discipline', "Unknown_Stream")
        department_name = latest_response.get('student_info', {}).get('branch', "Unknown_Department")
        semester = latest_response.get('student_info', {}).get('semester', "Unknown_Semester")
    
    # Replace spaces with underscores for the filename
    stream = stream.replace(" ", "_")
    department_name = department_name.replace(" ", "_")
    semester = semester.replace(" ", "_")
    subject_name_cleaned = subject_name.replace(" ", "_")

    # Define the columns for the Excel sheet
    columns = [
        "Student's Enrollment Number",
        "Subject",
        "Approximate % of total syllabus covered",
        "Regularity and punctuality in class",
        "Voice is audible and clear",
        "Use of teaching aid other than power point presentation",
        "Lecture preparation",
        "Lesson plan is being followed by teacher",
        "Depth of knowledge related to the course",
        "Reaction on students' comments or questions or in discussion",
        "Behavior with students",
        "Pace of teaching",
        "Effectiveness of communication with students",
        "Motivation provided to learn the course",
        "Explanation of the practical application of the course",
        "Clarity in answering to queries and explanation",
        "Rate (overall) the classes of the subject on following scale",
        "Clear explanation or demonstration of laboratory techniques, procedures",
        "Effective planning of laboratory activities",
        "Enforces to follow safety procedures and protocols in the laboratory"
    ]

    # Define the mapping from column names to JSON fields
    rating_fields = [
        "punctuality", "voice", "aids", "prep", "plan", "depth", "reaction",
        "behavior", "pace", "comm", "motiv", "expl", "clarity", "overall"
    ]
    lab_rating_fields = ["lab_expl", "lab_activ", "safety"]

    # Collect data rows
    data_rows = []
    for response in responses:
        for subject in response['subjects']:
            if subject['subject_name'] == subject_name:
                enrollment_number = response['student_info']['enrollment_number']
                subject_name_val = subject['subject_name']
                syllabus_covered = subject['syllabus_covered']
                ratings = subject['ratings']
                lab_ratings = subject['lab_ratings']

                # Convert syllabus_covered to integer (remove % if present)
                try:
                    syllabus_covered = int(syllabus_covered.replace('%', '')) if syllabus_covered else 0
                except (ValueError, AttributeError):
                    syllabus_covered = 0

                row = [
                    enrollment_number,
                    subject_name_val,
                    syllabus_covered
                ]
                # Convert ratings to integers
                for field in rating_fields:
                    rating_value = ratings.get(field, "")
                    try:
                        rating_value = int(rating_value) if rating_value else 0
                    except (ValueError, TypeError):
                        rating_value = 0
                    row.append(rating_value)
                # Convert lab ratings to integers
                for field in lab_rating_fields:
                    rating_value = lab_ratings.get(field, "")
                    try:
                        rating_value = int(rating_value) if rating_value else 0
                    except (ValueError, TypeError):
                        rating_value = 0
                    row.append(rating_value)
                data_rows.append(row)
                break

    # Create DataFrame
    df = pd.DataFrame(data_rows, columns=columns)

    # Generate Excel file with formatting
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write the DataFrame starting at row 4 (after the three new rows)
        df.to_excel(writer, index=False, sheet_name='Feedback', startrow=3)
        workbook = writer.book
        worksheet = writer.sheets['Feedback']

        # Add the two new rows above the heading
        # Row 1: UPL University of Sustainable Technology
        worksheet['A1'] = "UPL University of Sustainable Technology"
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        cell = worksheet['A1']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Times New Roman', size=20, bold=True)
        worksheet.row_dimensions[1].height = 30  # Adjust height for readability

        # Row 2: Shroff S. R. Rotary Institute of Chemical Technology
        worksheet['A2'] = "Shroff S. R. Rotary Institute of Chemical Technology"
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        cell = worksheet['A2']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Times New Roman', size=18, bold=True)
        worksheet.row_dimensions[2].height = 25  # Adjust height for readability

        # Row 3: Leave empty (spacer row)
        worksheet.row_dimensions[3].height = 15

        # Style header row (now row 4): center align and wrap text
        for cell in worksheet[4]:  # Heading row is now row 4
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))

        # Style data rows: center align and add borders
        for row in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row):  # Data starts at row 5
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))

        # Set fixed column widths using column letters
        for col_idx in range(len(columns)):
            column_letter = chr(65 + col_idx)  # A, B, C, ..., T (for 20 columns)
            if col_idx == 0:  # Enrollment Number column
                worksheet.column_dimensions[column_letter].width = 18.00
            elif col_idx == 1:  # Subject column
                worksheet.column_dimensions[column_letter].width = 40.00
            else:  # Other columns
                worksheet.column_dimensions[column_letter].width = 12.00

        # Set header row height (row 4)
        worksheet.row_dimensions[4].height = 100.00

    output.seek(0)

    # Create response
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={stream}_{department_name}_Semester_{semester}_{subject_name_cleaned}.xlsx'
    return response